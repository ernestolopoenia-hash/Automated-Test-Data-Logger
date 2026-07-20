"""
report.py
=========
ReportGenerator: produces every deliverable artifact for a completed
stress test run:

    * stress_test_results.csv   -- raw per-DUT data
    * stress_test_results.xlsx  -- formatted workbook version of the same
    * summary.json              -- machine-readable run summary
    * Stress_Test_Report.pdf    -- management-ready engineering report
"""

from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from config import AppConfig
from dut import DeviceUnderTest
from statistics_module import RunStatistics

CSV_FIELDS = [
    "timestamp", "test_id", "serial_number", "batch", "operator",
    "measured_voltage_v", "current_a", "power_w", "temperature_c",
    "humidity_pct", "leakage_current_ma", "voltage_ripple_mv",
    "voltage_spikes", "calibration_offset_v", "sensor_noise_v",
    "result", "failure_reason",
]


class ReportGenerator:
    """Coordinates every export format for a finished test run."""

    def __init__(self, config: AppConfig, duts: List[DeviceUnderTest], stats: RunStatistics):
        self.config = config
        self.duts = duts
        self.stats = stats
        os.makedirs(config.output_dir, exist_ok=True)

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def write_csv(self) -> str:
        path = os.path.join(self.config.output_dir, self.config.csv_filename)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
            writer.writeheader()
            for dut in self.duts:
                row = dut.to_row()
                writer.writerow({k: row.get(k, "") for k in CSV_FIELDS})
        return path

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def write_xlsx(self) -> str:
        path = os.path.join(self.config.output_dir, self.config.xlsx_filename)
        wb = Workbook()

        results_ws = wb.active
        results_ws.title = "Results"
        self._write_results_sheet(results_ws)

        summary_ws = wb.create_sheet("Summary")
        self._write_summary_sheet(summary_ws)

        wb.save(path)
        return path

    def _write_results_sheet(self, ws) -> None:
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="2C3E50")
        body_font = Font(name="Arial", size=10)
        pass_fill = PatternFill("solid", fgColor="E9F7EF")
        fail_fill = PatternFill("solid", fgColor="FDEDEC")

        headers = [
            "Timestamp", "Test ID", "Serial Number", "Batch", "Operator",
            "Voltage (V)", "Current (A)", "Power (W)", "Temp (C)",
            "Humidity (%)", "Leakage (mA)", "Ripple (mV)", "Spikes",
            "Cal. Offset (V)", "Sensor Noise (V)", "Result", "Failure Reason",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r, dut in enumerate(self.duts, start=2):
            row = dut.to_row()
            values = [
                row["timestamp"], row["test_id"], row["serial_number"],
                row["batch"], row["operator"], row["measured_voltage_v"],
                row["current_a"], row["power_w"], row["temperature_c"],
                row["humidity_pct"], row["leakage_current_ma"],
                row["voltage_ripple_mv"], row["voltage_spikes"],
                row["calibration_offset_v"], row["sensor_noise_v"],
                row["result"], row["failure_reason"],
            ]
            for c, value in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = body_font
                cell.fill = pass_fill if dut.passed else fail_fill

        widths = [19, 14, 12, 10, 14, 11, 10, 10, 9, 11, 11, 11, 8, 13, 14, 16, 22]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, ws) -> None:
        title_font = Font(name="Arial", bold=True, size=14, color="2C3E50")
        label_font = Font(name="Arial", bold=True, size=10)
        value_font = Font(name="Arial", size=10)

        ws["A1"] = f"{self.config.company.report_title} \u2014 Summary"
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")

        rows = [
            ("Company", self.config.company.company_name),
            ("Units Tested", self.stats.units_tested),
            ("Units Passed", self.stats.units_passed),
            ("Units Passed After Retest", self.stats.units_passed_after_retest),
            ("Units Failed", self.stats.units_failed),
            ("Pass Rate (%)", self.stats.pass_rate_pct),
            ("Fail Rate (%)", self.stats.fail_rate_pct),
            ("Average Voltage (V)", self.stats.avg_voltage_v),
            ("Average Current (A)", self.stats.avg_current_a),
            ("Average Temperature (C)", self.stats.avg_temperature_c),
            ("Average Leakage (mA)", self.stats.avg_leakage_ma),
            ("Maximum Temperature (C)", self.stats.max_temperature_c),
            ("Minimum Voltage (V)", self.stats.min_voltage_v),
            ("Total Test Time (s)", self.stats.total_test_time_s),
            ("Most Common Failure Mode", self.stats.most_common_failure_mode),
        ]
        for i, (label, value) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=label).font = label_font
            ws.cell(row=i, column=2, value=value).font = value_font

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 24

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def write_json(self) -> str:
        path = os.path.join(self.config.output_dir, self.config.json_filename)
        payload = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "company": self.config.company.company_name,
            "test_profile": self.config.company.report_title,
            "statistics": self.stats.as_dict(),
            "engineering_limits": {
                "voltage": {
                    "nominal_v": self.config.voltage.nominal_v,
                    "pass_min_v": self.config.voltage.pass_min_v,
                    "pass_max_v": self.config.voltage.pass_max_v,
                    "ripple_max_mv": self.config.voltage.ripple_max_mv,
                    "max_spikes": self.config.voltage.max_spikes,
                },
                "current": {
                    "nominal_a": self.config.current.nominal_a,
                    "tolerance_pct": self.config.current.tolerance_pct,
                    "min_a": round(self.config.current.min_a, 4),
                    "max_a": round(self.config.current.max_a, 4),
                },
                "thermal": {"max_temp_c": self.config.thermal.max_temp_c},
                "environment": {
                    "humidity_min_pct": self.config.environment.humidity_min_pct,
                    "humidity_max_pct": self.config.environment.humidity_max_pct,
                },
                "leakage": {"max_leakage_ma": self.config.leakage.max_leakage_ma},
            },
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)
        return path

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def write_pdf(self, plot_paths: List[str]) -> str:
        path = os.path.join(self.config.output_dir, self.config.pdf_filename)
        doc = SimpleDocTemplate(
            path, pagesize=letter,
            topMargin=0.75 * inch, bottomMargin=0.75 * inch,
            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
        )
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Title"], fontSize=24,
            textColor=colors.HexColor("#2C3E50"), spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle", parent=styles["Normal"], fontSize=13,
            textColor=colors.HexColor("#566573"), spaceAfter=4, alignment=1,
        )
        heading_style = ParagraphStyle(
            "SectionHeading", parent=styles["Heading1"], fontSize=15,
            textColor=colors.HexColor("#2C3E50"), spaceBefore=14, spaceAfter=8,
        )
        subheading_style = ParagraphStyle(
            "SubHeading", parent=styles["Heading2"], fontSize=12,
            textColor=colors.HexColor("#2980B9"), spaceBefore=10, spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "Body", parent=styles["Normal"], fontSize=10, leading=14,
        )

        story = []
        story += self._pdf_title_page(title_style, subtitle_style, body_style)
        story.append(PageBreak())
        story += self._pdf_executive_summary(heading_style, body_style)
        story += self._pdf_engineering_spec_table(heading_style)
        story.append(PageBreak())
        story += self._pdf_statistics_section(heading_style, body_style)
        story.append(PageBreak())
        story += self._pdf_plots_section(heading_style, subheading_style, plot_paths)
        story.append(PageBreak())
        story += self._pdf_failure_analysis(heading_style, subheading_style, body_style)
        story += self._pdf_conclusion(heading_style, body_style)

        doc.build(story, onFirstPage=self._draw_footer, onLaterPages=self._draw_footer)
        return path

    def _draw_footer(self, canvas, doc) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#95A5A6"))
        footer_text = (
            f"{self.config.company.company_name} \u2014 {self.config.company.lab_name} "
            f"| CONFIDENTIAL | Page {doc.page}"
        )
        canvas.drawCentredString(letter[0] / 2, 0.5 * inch, footer_text)
        canvas.restoreState()

    def _pdf_title_page(self, title_style, subtitle_style, body_style) -> list:
        elements = []
        elements.append(Spacer(1, 0.6 * inch))

        # Simple placeholder "logo": a drawn company monogram box built from
        # a table cell rather than an external image asset.
        logo_table = Table(
            [[self.config.company.company_name[:1]]], colWidths=[0.9 * inch], rowHeights=[0.9 * inch]
        )
        logo_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 28),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ]))
        centered_logo = Table([[logo_table]], colWidths=[6.5 * inch])
        centered_logo.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
        elements.append(centered_logo)

        elements.append(Spacer(1, 0.4 * inch))
        elements.append(Paragraph(self.config.company.company_name, title_style))
        elements.append(Paragraph(self.config.company.lab_name, subtitle_style))
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph(self.config.company.report_title, subtitle_style))
        elements.append(Spacer(1, 0.6 * inch))

        meta_rows = [
            ["Report Date:", datetime.now().strftime("%B %d, %Y")],
            ["Test Profile:", self.config.company.report_title],
            ["Units Tested:", str(self.stats.units_tested)],
            ["Overall Pass Rate:", f"{self.stats.pass_rate_pct}%"],
            ["Document Classification:", "Internal / Confidential"],
        ]
        meta_table = Table(meta_rows, colWidths=[2.2 * inch, 3.0 * inch])
        meta_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#2C3E50")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#D5D8DC")),
        ]))
        centered_meta = Table([[meta_table]], colWidths=[6.5 * inch])
        centered_meta.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
        elements.append(centered_meta)
        return elements

    def _pdf_executive_summary(self, heading_style, body_style) -> list:
        elements = [Paragraph("Executive Summary", heading_style)]
        disposition = "PASSED" if self.stats.fail_rate_pct < 50 else "REQUIRES REVIEW"
        summary_text = (
            f"A total of {self.stats.units_tested} devices under test (DUTs) were subjected to the "
            f"{self.config.company.report_title.lower()} at a nominal supply of "
            f"{self.config.voltage.nominal_v:.2f} V. "
            f"{self.stats.units_passed + self.stats.units_passed_after_retest} of "
            f"{self.stats.units_tested} units met all pass criteria "
            f"({self.stats.pass_rate_pct}% pass rate), while {self.stats.units_failed} units "
            f"({self.stats.fail_rate_pct}%) failed to meet one or more engineering specifications. "
            f"Of the passing units, {self.stats.units_passed_after_retest} required one automatic "
            f"retest before passing. The most frequently observed failure mode across the batch was "
            f"\u201c{self.stats.most_common_failure_mode}.\u201d Based on the aggregate pass rate, this "
            f"lot is assessed as <b>{disposition}</b> against the qualification criteria defined below."
        )
        elements.append(Paragraph(summary_text, body_style))
        elements.append(Spacer(1, 10))
        return elements

    def _pdf_engineering_spec_table(self, heading_style) -> list:
        elements = [Paragraph("Engineering Specifications &amp; Pass/Fail Criteria", heading_style)]
        v, c, t, e, lk = (
            self.config.voltage, self.config.current, self.config.thermal,
            self.config.environment, self.config.leakage,
        )
        data = [
            ["Parameter", "Specification"],
            ["Nominal Voltage", f"{v.nominal_v:.2f} V"],
            ["Voltage Pass Range", f"{v.pass_min_v:.2f} V \u2013 {v.pass_max_v:.2f} V"],
            ["Current", f"{c.nominal_a:.2f} A \u00b1 {c.tolerance_pct:.0f}% "
                        f"({c.min_a:.3f}\u2013{c.max_a:.3f} A)"],
            ["Temperature", f"Must remain below {t.max_temp_c:.0f}\u00b0C"],
            ["Leakage Current", f"Below {lk.max_leakage_ma:.1f} mA"],
            ["Voltage Ripple", f"Below {v.ripple_max_mv:.0f} mV"],
            ["Voltage Spikes", f"Maximum {v.max_spikes} spikes per test"],
            ["Humidity", f"{e.humidity_min_pct:.0f}% \u2013 {e.humidity_max_pct:.0f}% RH"],
        ]
        table = Table(data, colWidths=[2.2 * inch, 4.3 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F7")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10))
        return elements

    def _pdf_statistics_section(self, heading_style, body_style) -> list:
        elements = [Paragraph("Statistical Summary", heading_style)]
        s = self.stats
        data = [
            ["Metric", "Value"],
            ["Units Tested", str(s.units_tested)],
            ["Units Passed", str(s.units_passed)],
            ["Units Passed After Retest", str(s.units_passed_after_retest)],
            ["Units Failed", str(s.units_failed)],
            ["Pass Rate", f"{s.pass_rate_pct}%"],
            ["Fail Rate", f"{s.fail_rate_pct}%"],
            ["Average Voltage", f"{s.avg_voltage_v} V"],
            ["Average Current", f"{s.avg_current_a} A"],
            ["Average Temperature", f"{s.avg_temperature_c} \u00b0C"],
            ["Average Leakage Current", f"{s.avg_leakage_ma} mA"],
            ["Maximum Temperature Observed", f"{s.max_temperature_c} \u00b0C"],
            ["Minimum Voltage Observed", f"{s.min_voltage_v} V"],
            ["Total Cumulative Test Time", f"{s.total_test_time_s} s"],
            ["Average Test Time per DUT", f"{s.average_test_time_s} s"],
            ["Most Common Failure Mode", s.most_common_failure_mode],
        ]
        table = Table(data, colWidths=[3.0 * inch, 3.5 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2980B9")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F7")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        return elements

    def _pdf_plots_section(self, heading_style, subheading_style, plot_paths: List[str]) -> list:
        elements = [Paragraph("Distribution &amp; Trend Plots", heading_style)]
        titles = {
            "voltage_histogram.png": "Voltage Distribution",
            "temperature_histogram.png": "Temperature Distribution",
            "current_histogram.png": "Current Distribution",
            "leakage_histogram.png": "Leakage Current Distribution",
            "humidity_histogram.png": "Humidity Distribution",
            "pass_fail_pie.png": "Pass vs. Fail Distribution",
            "voltage_vs_temperature.png": "Voltage vs. Temperature",
            "current_vs_temperature.png": "Current vs. Temperature",
            "voltage_over_test_time.png": "Voltage Over Test Sequence",
            "failure_mode_bar_chart.png": "Failure Mode Frequency",
        }
        for path in plot_paths:
            filename = os.path.basename(path)
            title = titles.get(filename, filename)
            elements.append(Paragraph(title, subheading_style))
            elements.append(Image(path, width=5.6 * inch, height=3.6 * inch))
            elements.append(Spacer(1, 8))
        return elements

    def _pdf_failure_analysis(self, heading_style, subheading_style, body_style) -> list:
        elements = [Paragraph("Failure Analysis", heading_style)]
        if not self.stats.failure_mode_counts:
            elements.append(Paragraph(
                "No failures were recorded during this test run. All DUTs met specification.",
                body_style,
            ))
            return elements

        elements.append(Paragraph("Top Failure Modes", subheading_style))
        sorted_modes = sorted(
            self.stats.failure_mode_counts.items(), key=lambda kv: kv[1], reverse=True
        )
        data = [["Failure Mode", "Occurrences", "% of Failed Units"]]
        total_failed = max(self.stats.units_failed, 1)
        for mode, count in sorted_modes:
            data.append([mode, str(count), f"{round(100 * count / total_failed, 1)}%"])

        table = Table(data, colWidths=[3.2 * inch, 1.4 * inch, 1.9 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C0392B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FDEDEC")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D8DC")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10))
        return elements

    def _pdf_conclusion(self, heading_style, body_style) -> list:
        elements = [Paragraph("Engineering Conclusion &amp; Recommendations", heading_style)]
        s = self.stats
        if s.fail_rate_pct == 0:
            conclusion = (
                "All tested units met the defined voltage stress qualification criteria. "
                "No corrective action is required at this time. It is recommended to proceed "
                "with the next phase of qualification testing (e.g., Temperature Cycling or "
                "Power Cycling) using the same lot."
            )
        elif s.fail_rate_pct < 10:
            conclusion = (
                f"The observed fail rate of {s.fail_rate_pct}% is within typical first-pass "
                f"yield expectations for early qualification builds. The leading failure mode, "
                f"\u201c{s.most_common_failure_mode},\u201d should be root-caused before the lot is "
                f"released for volume production. It is recommended that Design and Test "
                f"Engineering review the affected units and, if a systemic cause is not identified, "
                f"proceed with cautious approval pending a confirmation run."
            )
        else:
            conclusion = (
                f"The observed fail rate of {s.fail_rate_pct}% exceeds acceptable qualification "
                f"thresholds. The dominant failure mode, \u201c{s.most_common_failure_mode},\u201d "
                f"warrants immediate root-cause investigation by Design, Process, and Reliability "
                f"Engineering. It is recommended that this lot be placed on hold pending "
                f"containment actions, and that a corrective action plan (CAPA) be opened before "
                f"any units are released downstream."
            )
        elements.append(Paragraph(conclusion, body_style))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(
            "This report was generated automatically by the Electronics Reliability Stress "
            "Test Simulator. All measurements shown are simulated and intended for software "
            "demonstration and portfolio purposes only.",
            ParagraphStyle("Disclaimer", parent=body_style, fontSize=8, textColor=colors.HexColor("#95A5A6")),
        ))
        return elements
